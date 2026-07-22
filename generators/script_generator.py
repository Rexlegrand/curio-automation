"""Génération du script JSON horodaté via Claude API (+ lecture Excel Type B).

Type B maths (ADDENDUM v2.6) : Claude classe aussi le sujet en image_route
(code_render / gpt_image) dans le même appel, sans coût API supplémentaire.
Le render_type et l'operation_data sont revérifiés en Python avant écriture
du script.json — aucun chiffre n'est fait confiance sans contrôle.
"""

import datetime
import json
import random
import re
import sys

import anthropic

from config import CLAUDE_MODEL, DATA_XLSX, ELEVENLABS_CONFIG, ENV

WORD_MIN, WORD_MAX = ELEVENLABS_CONFIG["word_count_target"]

SCHEMA_TYPE_A = """\
{
  "titre": "titre court pour la miniature (6 mots max)",
  "theme": "un parmi: sport, combat, velo, nature, histoire, maths, science, transport, meteo, default",
  "hook": "Attends... [fait surprenant en question ou affirmation choc]",
  "narration": "texte complet parlé, commence par la phrase hook, 85-100 mots au total",
  "segments": [
    {"start": 0, "end": 4, "role": "hook", "texte": "..."},
    {"start": 4, "end": 9, "role": "illustration_1", "texte": "..."},
    {"start": 9, "end": 13, "role": "curio_a", "texte": "..."},
    {"start": 13, "end": 18, "role": "illustration_2", "texte": "..."},
    {"start": 18, "end": 22, "role": "curio_b", "texte": "..."},
    {"start": 22, "end": 25, "role": "illustration_3", "texte": "..."},
    {"start": 25, "end": 30, "role": "cta", "texte": "..."}
  ],
  "illustrations": [
    {"description_visuelle": "description hyper-réaliste illustration 1 (photo style Wikipédia/magazine, jamais de personnage Curio, jamais de cartoon)"},
    {"description_visuelle": "illustration 2"},
    {"description_visuelle": "illustration 3"}
  ]
}"""

# ADDENDUM v2.6 §3 — classification image_route faite par Claude, même appel.
SCHEMA_TYPE_B_MATHS = """\
{
  "titre": "titre court pour la miniature (6 mots max)",
  "theme": "maths",
  "hook": "Attends... tu sais vraiment comment [compétence] ?",
  "narration": "texte complet parlé, commence par la phrase hook, 85-100 mots au total",
  "segments": [ même découpage 0-4/4-9/9-13/13-18/18-22/22-25/25-30 que Type A ],
  "image_route": "code_render" ou "gpt_image",
  "render_type": "division_posee | soustraction_colonnes | addition_colonnes | multiplication_posee | astuce_chaine, ou null si gpt_image",
  "operation_data": {
    "// si division_posee": "",
    "dividende": "int", "diviseur": "int (1 chiffre ; 2 chiffres autorisé seulement si niveau CM2)",
    "// si soustraction_colonnes ou addition_colonnes (clés identiques)": "",
    "nombre1": "int (nombre1 >= nombre2 pour la soustraction)", "nombre2": "int",
    "// si multiplication_posee": "",
    "multiplicande": "int", "multiplicateur": "int 1 chiffre (0-9)",
    "// si astuce_chaine (3 frames = 3 illustrations DIFFÉRENTES, jamais la même image répétée)": "",
    "titre": "ex: Multiplier par 5",
    "frames": [
      {"etapes": ["principe général, peut être en mots, pas de chiffre obligatoire"]},
      {"etapes": ["ligne 1 exemple 1 (ex: 46 × 5)", "ligne avec =", "dernière ligne = résultat avec ="]},
      {"etapes": ["ligne 1 exemple 2 (nombres DIFFÉRENTS de l'exemple 1)", "ligne avec =", "dernière ligne = résultat avec ="]}
    ],
    "// si gpt_image (null operation_data)": "ne pas inclure operation_data"
  },
  "illustrations": [
    {"description_visuelle": "SEULEMENT si image_route=gpt_image : diagramme pédagogique clair du concept (ex: axe de symétrie, parts de fraction, unité de mesure), fond cahier, pas de personnage Curio, pas de chiffre inventé"},
    { "illustration 2, même clé, vide [] si image_route=code_render": "..." },
    { "illustration 3, même clé": "..." }
  ]
}"""

SCHEMA_TYPE_B_FRANCAIS = """\
{
  "titre": "titre court pour la miniature (6 mots max)",
  "theme": "default",
  "hook": "Attends... tu sais vraiment comment [compétence] ?",
  "narration": "texte complet parlé, commence par la phrase hook, 85-100 mots au total",
  "segments": [ même découpage 0-4/4-9/9-13/13-18/18-22/22-25/25-30 que Type A ],
  "illustrations": [
    {
      "regle_exacte": "la règle de grammaire exacte (sert à la narration audio uniquement, jamais imprimée sur l'image)",
      "sujet_photo": "description photoréaliste CONCRÈTE du mot-exemple de CETTE illustration (ex: 'a wooden drum (tambour), close-up, natural light, on a wooden floor') — jamais une scène générique ou réutilisée d'un autre sujet, une par illustration",
      "mot_cle": "LE mot-exemple seul, correctement orthographié, sans phrase autour (ex: 'tambour') — c'est le SEUL texte dessiné sur l'image",
      "lettre_cle": "UNE seule lettre de mot_cle à colorer sur l'image pour la mettre en valeur (ex: 'b' pour tambour — la consonne qui impose la règle, ou la lettre de l'exception)",
      "exemple_correct": "phrase exemple correcte (sert à la narration audio uniquement)",
      "test_substitution_ok": "test de substitution qui marche (sert à la narration audio uniquement)",
      "conclusion_ok": "conclusion (sert à la narration audio uniquement)",
      "exemple_incorrect": "phrase exemple incorrecte (sert à la narration audio uniquement)",
      "test_substitution_ko": "test de substitution qui ne marche pas (sert à la narration audio uniquement)",
      "conclusion_ko": "conclusion (sert à la narration audio uniquement)"
    },
    { "illustration 2, mêmes clés, sujet_photo/mot_cle DIFFÉRENTS (mot-exemple différent)": "..." },
    { "illustration 3, mêmes clés, sujet_photo/mot_cle DIFFÉRENTS": "..." }
  ]
}"""

# ADDENDUM v2.6 §3 — règle de classification pour Claude (Type B maths uniquement)
CLASSIFICATION_RULES = """
Règle de classification image_route (obligatoire pour ce sujet maths) :
- Si le sujet implique une opération posée avec retenue/emprunt/potence
  → image_route = "code_render", choisir render_type parmi :
    division_posee | soustraction_colonnes | addition_colonnes | multiplication_posee
  → remplir operation_data avec les clés exactes de ce render_type (voir schéma),
    illustrations = [] (tableau vide, non utilisé pour code_render)
- Si le sujet est une astuce de calcul mental présentable comme une chaîne
  d'égalités (ex: ×5 = ×10÷2, ×4 = ×2×2)
  → image_route = "code_render", render_type = "astuce_chaine"
  → operation_data = {"titre": ..., "frames": [frame_principe, frame_exemple1, frame_exemple2]}
    Ce sont 3 illustrations DIFFÉRENTES (jamais la même image répétée 3 fois) :
    - frame 1 (principe) : la règle générale. Peut être en mots (ex: "Ajouter 9
      = Ajouter 10 puis retirer 1"), pas besoin d'un nombre concret ici.
    - frame 2 (exemple 1) et frame 3 (exemple 2) : DEUX exemples chiffrés
      DIFFÉRENTS, entièrement résolus — c'est ce qui prouve que l'astuce marche
      à chaque fois. Si la compétence source mentionne "2 exemples" ou
      équivalent, respecte bien ce nombre.
    - FORMAT STRICT des frames 2 et 3 (chaque ligne vérifiée automatiquement) :
      la première ligne est une expression seule, SANS signe = (ex: "7 + 9") ;
      chaque ligne suivante est une égalité COMPLÈTE "valeur = valeur" avec un
      nombre explicite des DEUX côtés (ex: "7 + 10 = 17", PAS "= 17") ; jamais
      de ligne commençant par "=", jamais de "?", jamais de variable.
      UNE LIGNE = UNE SEULE OPÉRATION = UN SEUL RÉSULTAT VISIBLE. Ne jamais
      fusionner deux étapes sur la même ligne (ex INTERDIT : "7 + 9 = 7 + 10 - 1"
      cache le résultat intermédiaire 17). Décompose plutôt en 3 lignes :
      "7 + 9" / "7 + 10 = 17" / "17 - 1 = 16" — chaque ligne réutilise le
      résultat de la précédente et n'applique qu'un seul calcul.
- Sinon (notion, concept visuel, règle sans calcul chiffré : symétrie,
  fractions en parts, unités de mesure)
  → image_route = "gpt_image", render_type = null, ne pas inclure operation_data,
  → remplir illustrations avec 3 description_visuelle (diagramme pédagogique)
Vérifie toi-même chaque valeur numérique d'operation_data avant de répondre :
le résultat doit être mathématiquement exact.
"""


def pick_competence(niveau, matiere):
    """Choisit une compétence dans data/Competences_Curio.xlsx (onglet = niveau).

    L'en-tête (Matière | Difficulté | Compétence) n'est pas forcément en ligne 1
    (l'onglet CP a des lignes de statut au-dessus) : on filtre sur le contenu.
    """
    if not DATA_XLSX.exists():
        sys.exit(f"Erreur : {DATA_XLSX} introuvable. Dépose le fichier Excel puis relance.")
    import openpyxl

    wb = openpyxl.load_workbook(DATA_XLSX, read_only=True)
    if niveau not in wb.sheetnames:
        sys.exit(f"Erreur : onglet '{niveau}' absent du Excel. Onglets : {wb.sheetnames}")
    ws = wb[niveau]
    candidates = []
    for r in ws.iter_rows(values_only=True):
        if not r or len(r) < 3 or not r[0] or not r[2]:
            continue
        col_matiere = str(r[0]).strip().lower()
        if col_matiere == "matière":
            continue
        if matiere.strip().lower() in col_matiere:
            candidates.append(r)
    if not candidates:
        sys.exit(f"Erreur : aucune compétence '{matiere}' dans l'onglet {niveau}.")
    row = random.choice(candidates)
    competence = {"matiere": str(row[0]), "difficulte": str(row[1]), "competence": str(row[2])}
    print(f"Compétence choisie : [{niveau}] {competence['matiere']} — {competence['competence']} (difficulté {competence['difficulte']})")
    return competence


CTA_INSTRUCTIONS = {
    "abonnement": (
        "Le dernier segment (role cta) est un appel à s'abonner, court et naturel "
        "(variante de : Abonne-toi pour une nouvelle curiosité chaque jour !)."
    ),
    "commentaire": (
        "Le dernier segment (role cta) demande de commenter le mot CURIO (toujours CURIO, "
        "jamais un autre mot) pour recevoir une activité pédagogique gratuite "
        "(ex : Commente CURIO et je t'envoie une activité gratuite !). "
        "Ajoute aussi un champ \"cta_mot\": \"CURIO\" dans le JSON."
    ),
}


def _build_prompt(reel_type, sujet, niveau, matiere, cta_type):
    is_maths = reel_type == "competence" and matiere and "math" in matiere.lower()
    if reel_type == "curiosite":
        schema = SCHEMA_TYPE_A
        contexte = f"Sujet du Reel (Type A — Curiosité du jour) : {sujet}"
        regles = (
            "- Les descriptions d'illustrations doivent être 100% photoréalistes, "
            "comme des photos Wikipédia ou de magazine. Jamais de personnage Curio, jamais de cartoon.\n"
            "- Si un schéma est nécessaire : flèches + chiffres simples, minimaliste."
        )
    elif is_maths:
        schema = SCHEMA_TYPE_B_MATHS
        contexte = f"Sujet du Reel (Type B — Compétence maths, niveau {niveau}) : {sujet}"
        regles = (
            "- EXACTITUDE PÉDAGOGIQUE STRICTE : chaque chiffre mathématiquement correct, "
            "méthode Éducation Nationale uniquement. Vérifie chaque calcul deux fois.\n"
            + CLASSIFICATION_RULES
        )
    else:
        schema = SCHEMA_TYPE_B_FRANCAIS
        contexte = f"Sujet du Reel (Type B — Compétence français, niveau {niveau}) : {sujet}"
        regles = (
            "- EXACTITUDE PÉDAGOGIQUE STRICTE : chaque mot correctement orthographié, "
            "tous les accents présents, règle conforme aux programmes officiels."
        )
    schema = schema.replace("85-100", f"{WORD_MIN}-{WORD_MAX}")
    cta_instruction = CTA_INSTRUCTIONS[cta_type]
    return f"""Tu écris le script d'un Reel Instagram de 28-35 secondes pour @curio.education,
compte éducatif français pour enfants de primaire (CP-CM2) et leurs parents.
Le narrateur est Curio, un pingouin curieux et enthousiaste. Ton : simple, vivant,
phrases courtes, vocabulaire accessible à un enfant de 8 ans.

{contexte}

Contraintes :
- La narration fait STRICTEMENT entre {WORD_MIN} et {WORD_MAX} mots (la voix lit plus
  lentement que prévu — 141 à 160 mots/minute mesurés en production, jamais 180 : cette
  fourchette est calibrée pour rester entre 28 et 35 secondes, jamais plus de 35, même au
  débit le plus lent observé).
- La narration commence par la phrase hook exacte.
- Les segments suivent le découpage timecode imposé et la somme des textes = la narration.
- AUCUNE DIGRESSION : chaque phrase sert directement le sujet principal. Si une info est
  seulement cousine du sujet (autre règle, autre récompense, anecdote hors sujet), ne pas l'inclure.
- {cta_instruction}
{regles}

Réponds UNIQUEMENT avec un objet JSON valide (aucun texte autour, pas de bloc markdown) suivant ce schéma :
{schema}"""


def _count_words(text):
    return len(text.split())


_SAFE_EXPR = re.compile(r"^[0-9+\-*/(). ]+$")
_OPERATOR = re.compile(r"[+\-*/]")


def _count_operators(expr):
    """Compte les opérateurs d'une expression (ignore un '-' de signe en tête)."""
    normalized = expr.replace("×", "*").replace("x", "*").replace("X", "*").replace("÷", "/").strip()
    if normalized.startswith("-"):
        normalized = normalized[1:]
    return len(_OPERATOR.findall(normalized))


def _safe_eval_arithmetic(expr):
    """Évalue une expression arithmétique simple (chiffres/+-*/()) sans exec de code arbitraire."""
    normalized = expr.replace("×", "*").replace("x", "*").replace("X", "*").replace("÷", "/").replace(",", ".")
    normalized = normalized.strip()
    if not _SAFE_EXPR.match(normalized):
        raise ValueError(f"expression non numérique : {expr!r}")
    return eval(normalized, {"__builtins__": {}}, {})  # noqa: S307 — charset restreint à [0-9+-*/(). ] ci-dessus


def _validate_astuce_chaine(operation_data):
    """Vérifie les 3 frames (principe/exemple 1/exemple 2) de l'astuce.

    Frame 0 (principe) peut contenir du texte non chiffré (règle en mots) —
    une ligne non numérique y est ignorée, rien à vérifier. Frames 1 et 2
    (exemples) doivent être entièrement chiffrés : chaque ligne 'a = b' est
    vérifiée arithmétiquement, chaque ligne ne porte qu'une seule opération
    (jamais deux étapes fusionnées, sinon le résultat intermédiaire disparaît),
    et les deux exemples doivent porter sur des nombres différents (c'est ce
    qui prouve que l'astuce marche à chaque fois).
    """
    frames = operation_data.get("frames")
    if not frames or not isinstance(frames, list) or len(frames) != 3:
        raise ValueError("astuce_chaine : operation_data.frames doit contenir exactement 3 frames (principe, exemple 1, exemple 2)")

    for i, frame in enumerate(frames):
        etapes = frame.get("etapes") if isinstance(frame, dict) else None
        if not etapes or not isinstance(etapes, list):
            raise ValueError(f"astuce_chaine : frame {i + 1} sans etapes valides")
        strict = i > 0  # frame 0 = principe, peut contenir du texte non chiffré
        for ligne in etapes:
            if "=" not in ligne:
                if strict and _count_operators(ligne) != 1:
                    raise ValueError(
                        f"astuce_chaine : frame {i + 1} première ligne '{ligne}' doit être une "
                        "expression à une seule opération (ex: '7 + 9')"
                    )
                continue
            gauche, droite = ligne.rsplit("=", 1)
            try:
                valeur_gauche = _safe_eval_arithmetic(gauche)
                valeur_droite = _safe_eval_arithmetic(droite)
            except (ValueError, SyntaxError, ZeroDivisionError) as exc:
                if strict:
                    raise ValueError(f"astuce_chaine : frame {i + 1} ligne illisible '{ligne}' ({exc})") from exc
                continue
            if abs(valeur_gauche - valeur_droite) > 1e-6:
                raise ValueError(f"astuce_chaine : frame {i + 1} ligne fausse '{ligne}' ({valeur_gauche} != {valeur_droite})")
            if strict and (_count_operators(gauche) > 1 or _count_operators(droite) > 1):
                raise ValueError(
                    f"astuce_chaine : frame {i + 1} ligne '{ligne}' fusionne plusieurs opérations "
                    "sur une même ligne (une ligne = un seul calcul, résultat intermédiaire visible)"
                )

    if frames[1]["etapes"][0].strip() == frames[2]["etapes"][0].strip():
        raise ValueError("astuce_chaine : exemple 1 et exemple 2 utilisent le même nombre, il en faut deux différents")


def _validate_operation_data(render_type, operation_data):
    """Vérifie les opérandes fournis par Claude avant toute écriture (ADDENDUM v2.6 §3).

    Pour les opérations posées, le résultat est de toute façon recalculé par le
    renderer (jamais celui de Claude) : on vérifie seulement le TYPE et les
    contraintes des opérandes. Pour astuce_chaine, Claude fournit du texte
    libre : on vérifie l'exactitude arithmétique de chaque ligne.
    """
    if operation_data is None:
        raise ValueError(f"{render_type} : operation_data manquant")

    if render_type == "division_posee":
        dividende, diviseur = operation_data.get("dividende"), operation_data.get("diviseur")
        if not isinstance(dividende, int) or not isinstance(diviseur, int):
            raise ValueError("division_posee : dividende/diviseur doivent être des entiers")
        if diviseur == 0:
            raise ValueError("division_posee : diviseur ne peut pas être 0")
    elif render_type in ("soustraction_colonnes", "addition_colonnes"):
        n1, n2 = operation_data.get("nombre1"), operation_data.get("nombre2")
        if not isinstance(n1, int) or not isinstance(n2, int):
            raise ValueError(f"{render_type} : nombre1/nombre2 doivent être des entiers")
        if render_type == "soustraction_colonnes" and n1 < n2:
            raise ValueError(f"soustraction_colonnes : nombre1 ({n1}) < nombre2 ({n2}), résultat négatif")
    elif render_type == "multiplication_posee":
        multiplicande, multiplicateur = operation_data.get("multiplicande"), operation_data.get("multiplicateur")
        if not isinstance(multiplicande, int) or not isinstance(multiplicateur, int):
            raise ValueError("multiplication_posee : multiplicande/multiplicateur doivent être des entiers")
        if not (0 <= multiplicateur <= 9):
            raise ValueError(f"multiplication_posee : multiplicateur {multiplicateur} doit être à 1 chiffre")
    elif render_type == "astuce_chaine":
        _validate_astuce_chaine(operation_data)
    else:
        raise ValueError(f"render_type inconnu : {render_type}")


def _validate_classification(script):
    """Type B maths uniquement : vérifie image_route/render_type/operation_data."""
    route = script.get("image_route")
    if route not in ("code_render", "gpt_image"):
        raise ValueError(f"image_route invalide ou manquant : {route!r}")
    if route == "code_render":
        _validate_operation_data(script.get("render_type"), script.get("operation_data"))
    elif route == "gpt_image":
        illustrations = script.get("illustrations")
        if not illustrations or len(illustrations) != 3:
            raise ValueError("gpt_image : 3 illustrations (description_visuelle) attendues")


REQUIRED_FRANCAIS_KEYS = [
    "regle_exacte", "sujet_photo", "mot_cle", "lettre_cle", "exemple_correct",
    "test_substitution_ok", "conclusion_ok", "exemple_incorrect",
    "test_substitution_ko", "conclusion_ko",
]


def _validate_francais_illustrations(script):
    """Type B français : chaque illustration doit avoir un sujet_photo + mot_cle/lettre_cle
    concrets et distincts (bug v2.7 : sans sujet_photo, GPT Image pioche du contenu générique
    dans les références ; bug v2.9 : mot_cle/lettre_cle remplacent le paragraphe complet
    illisible sur mobile — seul un mot court avec 1 lettre colorée est dessiné sur l'image)."""
    illustrations = script.get("illustrations")
    if not illustrations or len(illustrations) != 3:
        raise ValueError("français : 3 illustrations attendues")
    sujets = []
    for i, illus in enumerate(illustrations):
        missing = [k for k in REQUIRED_FRANCAIS_KEYS if not illus.get(k)]
        if missing:
            raise ValueError(f"français : illustration {i + 1} sans {', '.join(missing)}")
        lettre_cle = illus["lettre_cle"].strip()
        if len(lettre_cle) != 1:
            raise ValueError(f"français : illustration {i + 1} lettre_cle doit être une seule lettre (reçu '{lettre_cle}')")
        if lettre_cle.lower() not in illus["mot_cle"].lower():
            raise ValueError(f"français : illustration {i + 1} lettre_cle '{lettre_cle}' absente de mot_cle '{illus['mot_cle']}'")
        sujets.append(illus["sujet_photo"].strip().lower())
    if len(set(sujets)) != len(sujets):
        raise ValueError("français : sujet_photo identique/répété sur plusieurs illustrations, il en faut un différent par mot-exemple")


def generate_script(reel_type, sujet, niveau=None, matiere=None, cta_type="abonnement"):
    """Appelle Claude, valide le nombre de mots (+ classification maths), retourne le dict script."""
    is_maths = reel_type == "competence" and matiere and "math" in matiere.lower()
    client = anthropic.Anthropic(api_key=ENV["ANTHROPIC_API_KEY"])
    prompt = _build_prompt(reel_type, sujet, niveau, matiere, cta_type)
    feedback = ""
    for attempt in range(3):
        response = client.messages.create(
            model=CLAUDE_MODEL,
            max_tokens=8000,
            messages=[{"role": "user", "content": prompt + feedback}],
        )
        raw = "".join(b.text for b in response.content if b.type == "text").strip()
        if "{" not in raw:
            raise ValueError(f"Réponse Claude sans JSON : {raw[:200]}")
        raw = raw[raw.index("{"):raw.rindex("}") + 1]
        script = json.loads(raw)
        wc = _count_words(script["narration"])

        problems = []
        if not (WORD_MIN <= wc <= WORD_MAX):
            problems.append(f"narration à {wc} mots (cible {WORD_MIN}-{WORD_MAX})")
        if is_maths:
            try:
                _validate_classification(script)
            except ValueError as exc:
                problems.append(str(exc))
        elif reel_type == "competence":
            try:
                _validate_francais_illustrations(script)
            except ValueError as exc:
                problems.append(str(exc))

        if not problems:
            break
        print(f"Script invalide ({'; '.join(problems)}), régénération...")
        feedback = (
            f"\n\nTa réponse précédente était invalide : {'; '.join(problems)}. "
            "Corrige et réponds à nouveau avec le JSON complet en respectant strictement le schéma."
        )
    else:
        raise ValueError(f"Script invalide après 3 tentatives : {'; '.join(problems)}")

    if not is_maths:
        script.setdefault("image_route", "gpt_image")
        script.setdefault("render_type", None)
        script.setdefault("operation_data", None)

    script["genere_le"] = datetime.datetime.now().isoformat(timespec="seconds")
    script["type"] = reel_type
    script["sujet"] = sujet
    script["niveau"] = niveau
    script["matiere"] = matiere
    script["cta_type"] = cta_type
    script["word_count"] = wc
    return script


def save_script(script, output_dir):
    path = output_dir / "script.json"
    path.write_text(json.dumps(script, ensure_ascii=False, indent=2))
    return path
